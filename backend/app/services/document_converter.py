"""Server-side conversion of binary documents to Markdown.

Claude Code CLI's `Read` tool rejects binary files, so uploaded Excel/Word/PDF
specs are unreadable inside the task container. This module extracts their text
content into a Markdown sibling file (`{stored_path}.md`) on the host upload
volume. The `.md` file is copied into `/workspace/uploads/` together with the
original by `DockerService.copy_uploads_to_container()`, and
`ClaudeCodeService._prepare_uploads()` points Claude at the readable version.

All functions are synchronous (openpyxl/python-docx/pdfplumber are sync);
callers on the async path run them via `asyncio.to_thread`.
"""
import logging
import os
from typing import List, Optional

logger = logging.getLogger(__name__)

# Extensions we can extract to Markdown. Anything else is either directly
# readable by Claude Code (text, Markdown, images) or unsupported.
CONVERTIBLE_EXTENSIONS = {".xlsx", ".xlsm", ".docx", ".pdf"}

CONVERTED_SUFFIX = ".md"


def is_convertible(filename: str) -> bool:
    """Whether the file is a binary document we can convert to Markdown."""
    return os.path.splitext(filename)[1].lower() in CONVERTIBLE_EXTENSIONS


def converted_path_for(stored_path: str) -> str:
    """Path of the Markdown conversion sitting next to the stored binary."""
    return stored_path + CONVERTED_SUFFIX


def ensure_converted(stored_path: str) -> Optional[str]:
    """
    Make sure `{stored_path}.md` exists and is up to date for a convertible
    binary. Returns the converted path, or None when the file is not
    convertible or conversion failed (the upload itself is never rejected).
    """
    if not is_convertible(stored_path) or not os.path.isfile(stored_path):
        return None

    out_path = converted_path_for(stored_path)
    try:
        if (
            os.path.isfile(out_path)
            and os.path.getmtime(out_path) >= os.path.getmtime(stored_path)
        ):
            return out_path

        markdown = convert_to_markdown(stored_path)
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(markdown)
        return out_path
    except Exception:
        logger.warning("Document conversion failed for %s", stored_path, exc_info=True)
        # Remove a stale/partial conversion so Claude is not fed old content.
        try:
            if os.path.isfile(out_path):
                os.remove(out_path)
        except OSError:
            pass
        return None


def convert_to_markdown(path: str) -> str:
    """Extract a binary document's content as Markdown. Raises on failure."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        body = _xlsx_to_markdown(path)
    elif ext == ".docx":
        body = _docx_to_markdown(path)
    elif ext == ".pdf":
        body = _pdf_to_markdown(path)
    else:
        raise ValueError(f"Unsupported extension: {ext}")

    header = (
        f"<!-- Converted from {os.path.basename(path)} by Xolvien. "
        "Layout/styling is not preserved; content only. -->\n\n"
    )
    return header + body


# ── Excel ──────────────────────────────────────────────────────────────────────

def _cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    # Keep the Markdown table on one line per row.
    return text.replace("|", "\\|").replace("\n", "<br>")


def _markdown_table(rows: List[List[str]]) -> str:
    width = max(len(r) for r in rows)
    padded = [r + [""] * (width - len(r)) for r in rows]
    table = ["| " + " | ".join(r) + " |" for r in padded]
    # The first row acts as the header for readability.
    table.insert(1, "|" + " --- |" * width)
    return "\n".join(table) + "\n"


def _xlsx_to_markdown(path: str) -> str:
    from openpyxl import load_workbook

    # data_only=True reads cached formula results instead of formula strings.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        parts: List[str] = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [_cell_text(v) for v in row]
                if any(c.strip() for c in cells):
                    rows.append(cells)
            parts.append(f"## Sheet: {ws.title}\n")
            parts.append(_markdown_table(rows) if rows else "(empty sheet)\n")
        return "\n".join(parts)
    finally:
        wb.close()


# ── Word ───────────────────────────────────────────────────────────────────────

def _docx_to_markdown(path: str) -> str:
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = Document(path)
    parts: List[str] = []

    # iter_inner_content() preserves the original paragraph/table order.
    for block in doc.iter_inner_content():
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            style = (block.style.name or "") if block.style else ""
            if style.startswith("Heading"):
                try:
                    level = min(int(style.split()[-1]), 6)
                except ValueError:
                    level = 2
                parts.append(f"{'#' * level} {text}\n")
            elif style.startswith("List"):
                parts.append(f"- {text}")
            else:
                parts.append(f"{text}\n")
        elif isinstance(block, Table):
            rows = [
                [_cell_text(cell.text) for cell in row.cells]
                for row in block.rows
            ]
            if rows:
                parts.append(_markdown_table(rows))

    return "\n".join(parts)


# ── PDF ────────────────────────────────────────────────────────────────────────

def _pdf_to_markdown(path: str) -> str:
    import pdfplumber

    parts: List[str] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            parts.append(f"## Page {i}\n")
            text = (page.extract_text() or "").strip()
            parts.append(text + "\n" if text else "(no extractable text)\n")
            for table in page.extract_tables():
                rows = [[_cell_text(c) for c in row] for row in table if row]
                if rows:
                    parts.append(_markdown_table(rows))
    return "\n".join(parts)
