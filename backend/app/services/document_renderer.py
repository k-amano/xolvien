"""Document YAML (v1.1) -> HTML / Excel renderer (Sprint 3 step 3.3).

One generic renderer for every doc type, implementing the rendering contract
of docs/document-format.md section 6. The HTML renderer is adapted from the
user-provided prototype (docs/upload/renderer.py) and keeps its layout,
styles, and cell-merge semantics; on top of it this version localizes the
fixed strings (ja/en) and embeds images as base64 data URIs resolved from the
per-task asset snapshot (documents/tasks/{task_id}/assets/).

The Excel renderer walks the same tree with openpyxl: numbered headings,
merged cells (colspan/rowspan), multi-row headers, row-header styling, page
breaks, embedded images, and preformatted Mermaid/code blocks.

The page frame is built in code for now; externalizing it into user-editable
templates is step 3.5.
"""
import base64
import html
import mimetypes
import os
from datetime import datetime
from io import BytesIO
from typing import Any, List, Optional

I18N = {
    "ja": {
        "table": "表",
        "figure": "図",
        "revision_history": "改訂履歴",
        "version": "版数",
        "date": "日付",
        "author": "作成者",
        "summary": "改訂概要",
        "image_missing": "（画像なし）",
        "cover_version": "第 {version} 版",
        "cover_organization": "組織",
        "cover_department": "部署",
        "cover_author": "作成者",
        "cover_reviewers": "査閲者",
        "cover_approver": "承認者",
        "generated_at": "生成日時",
    },
    "en": {
        "table": "Table",
        "figure": "Figure",
        "revision_history": "Revision History",
        "version": "Version",
        "date": "Date",
        "author": "Author",
        "summary": "Summary",
        "image_missing": "(image missing)",
        "cover_version": "Version {version}",
        "cover_organization": "Organization",
        "cover_department": "Department",
        "cover_author": "Author",
        "cover_reviewers": "Reviewers",
        "cover_approver": "Approver",
        "generated_at": "Generated at",
    },
}

_COVER_META_FIELDS = [
    ("organization", "cover_organization"),
    ("department", "cover_department"),
    ("author", "cover_author"),
    ("reviewers", "cover_reviewers"),
    ("approver", "cover_approver"),
]


def _cell_parts(cell: Any) -> tuple[Any, int, int]:
    """(value, colspan, rowspan) of a scalar-or-merge-object cell."""
    if isinstance(cell, dict):
        return cell.get("value"), cell.get("colspan", 1), cell.get("rowspan", 1)
    return cell, 1, 1


def _cell_text(value: Any) -> str:
    return "" if value is None else str(value)


# ── HTML ───────────────────────────────────────────────────────────────────────

_HTML_STYLE = """\
        @page {
            size: A4;
            margin: 20mm 15mm 20mm 15mm;
        }

        :root {
            --font-gothic: "游ゴシック", "Yu Gothic", "YuGothic", "ＭＳ ゴシック", "MS Gothic", sans-serif;
            --font-mincho: "游明朝", "Yu Mincho", "YuMincho", "ＭＳ 明朝", "MS Mincho", serif;
        }

        body {
            font-family: var(--font-mincho);
            font-size: 10.5pt;
            line-height: 1.5;
            color: #000;
            margin: 0;
            padding: 20px;
            max-width: 210mm;
        }

        .header {
            font-family: var(--font-gothic);
            font-size: 9pt;
            border-bottom: 1pt solid #000;
            padding-bottom: 3pt;
            margin-bottom: 15pt;
        }

        .cover {
            font-family: var(--font-gothic);
            page-break-after: always;
            text-align: center;
            padding-top: 100pt;
            min-height: 500pt;
        }
        .cover-title {
            font-size: 24pt;
            font-weight: bold;
            color: #00876c;
            margin-bottom: 15pt;
            line-height: 1.4;
        }
        .cover-subtitle {
            font-size: 14pt;
            margin-bottom: 40pt;
        }
        .cover-info {
            font-size: 12pt;
            margin: 8pt 0;
        }
        .cover-meta {
            margin-top: 60pt;
            font-size: 11pt;
        }
        .cover-meta table {
            margin: 0 auto;
            border-collapse: collapse;
        }
        .cover-meta th, .cover-meta td {
            padding: 4pt 12pt;
            text-align: left;
            border: none;
        }
        .cover-meta th {
            background: none;
            font-weight: normal;
        }

        .revision-page {
            page-break-after: always;
        }
        .revision-title {
            font-family: var(--font-gothic);
            font-size: 14pt;
            font-weight: bold;
            text-align: center;
            margin-bottom: 20pt;
        }

        h1 {
            font-family: var(--font-gothic);
            font-size: 14pt;
            font-weight: bold;
            margin: 20pt 0 10pt 0;
            padding: 6pt 10pt;
            background-color: #00876c;
            color: #fff;
        }

        h2 {
            font-family: var(--font-gothic);
            font-size: 12pt;
            font-weight: bold;
            margin: 15pt 0 8pt 0;
            padding: 4pt 0 4pt 10pt;
            border-left: 4pt solid #2e75b6;
            background-color: #f0f0f0;
        }

        h3 {
            font-family: var(--font-gothic);
            font-size: 11pt;
            font-weight: bold;
            margin: 12pt 0 6pt 0;
        }

        p {
            margin: 6pt 0;
            text-align: justify;
        }

        .table-caption {
            font-family: var(--font-gothic);
            font-size: 10pt;
            font-weight: bold;
            margin: 10pt 0 4pt 0;
        }

        .figure-caption {
            font-family: var(--font-gothic);
            font-size: 10pt;
            margin: 4pt 0 10pt 0;
        }

        table {
            font-family: var(--font-gothic);
            width: 100%;
            border-collapse: collapse;
            margin: 0 0 10pt 0;
            font-size: 10pt;
        }
        th, td {
            border: 1pt solid #000;
            padding: 4pt 6pt;
            text-align: left;
            vertical-align: top;
        }
        th {
            background-color: #d9d9d9;
            font-weight: normal;
            text-align: center;
        }
        .row-header {
            background-color: #d9d9d9;
            text-align: center;
        }

        ul, ol {
            font-family: var(--font-gothic);
            margin: 6pt 0 6pt 25pt;
            padding: 0;
        }
        li {
            margin: 3pt 0;
        }

        .mermaid-container {
            margin: 10pt 0;
            text-align: center;
        }
        .mermaid {
            display: inline-block;
        }

        .code-block {
            background-color: #f5f5f5;
            border: 1pt solid #ccc;
            padding: 8pt 10pt;
            margin: 8pt 0;
            font-family: "ＭＳ ゴシック", "MS Gothic", monospace;
            font-size: 9pt;
            white-space: pre-wrap;
            overflow-x: auto;
        }
        .code-caption {
            font-size: 10pt;
            margin: 8pt 0 4pt 0;
        }

        .note {
            margin: 10pt 0;
            padding: 8pt 12pt;
            border-left: 4pt solid #666;
            background-color: #f9f9f9;
        }
        .note-info {
            border-left-color: #2e75b6;
            background-color: #e8f4fc;
        }
        .note-warning {
            border-left-color: #e6a700;
            background-color: #fff8e6;
        }
        .note-important {
            border-left-color: #c00;
            background-color: #fee;
        }

        .image-container {
            margin: 10pt 0;
        }
        .image-container.align-center {
            text-align: center;
        }
        .image-container.align-right {
            text-align: right;
        }
        .image-container img {
            max-width: 100%;
        }
        .image-missing {
            display: inline-block;
            padding: 20pt 40pt;
            border: 1pt dashed #999;
            color: #666;
            background-color: #f5f5f5;
        }

        .page-break {
            page-break-before: always;
        }

        .footer {
            margin-top: 30pt;
            font-size: 9pt;
            text-align: right;
            color: #666;
            border-top: 1pt solid #ccc;
            padding-top: 5pt;
        }
"""


class HtmlDocumentRenderer:
    """Renders a parsed document dict to a self-contained HTML string."""

    def __init__(self, assets_dir: Optional[str] = None, generated_at: Optional[datetime] = None):
        self.assets_dir = assets_dir
        self.generated_at = generated_at
        self.table_count = 0
        self.figure_count = 0
        self.t = I18N["ja"]

    def render(self, doc: dict) -> str:
        self.table_count = 0
        self.figure_count = 0
        lang = doc.get("language", "ja")
        self.t = I18N.get(lang, I18N["ja"])

        parts = [self._head(doc), "<body>"]
        parts.append(self._cover(doc))
        if doc.get("revisions"):
            parts.append(self._revisions(doc))
        parts.append(f'<div class="header">{html.escape(doc["title"])}</div>')
        for i, section in enumerate(doc.get("sections", []), 1):
            parts.append(self._section(section, [i]))
        stamp = (self.generated_at or datetime.now()).strftime("%Y-%m-%d %H:%M")
        parts.append(f'<div class="footer">{self.t["generated_at"]}: {stamp}</div>')
        parts.append("</body></html>")
        return "\n".join(parts)

    def _head(self, doc: dict) -> str:
        title = html.escape(doc.get("title", ""))
        return (
            "<!DOCTYPE html>\n"
            f'<html lang="{doc.get("language", "ja")}">\n'
            "<head>\n"
            '    <meta charset="UTF-8">\n'
            f"    <title>{title}</title>\n"
            f"    <style>\n{_HTML_STYLE}    </style>\n"
            '    <script src="https://cdn.jsdelivr.net/npm/mermaid/dist/mermaid.min.js"></script>\n'
            "    <script>mermaid.initialize({startOnLoad: true});</script>\n"
            "</head>"
        )

    def _multiline(self, text: str) -> str:
        return html.escape(text).replace("\n", "<br>")

    def _cover(self, doc: dict) -> str:
        cover = doc.get("cover", {})
        parts = ['<div class="cover">']
        parts.append(f'<div class="cover-title">{self._multiline(doc.get("title", ""))}</div>')
        if "subtitle" in cover:
            parts.append(f'<div class="cover-subtitle">{html.escape(cover["subtitle"])}</div>')
        if "version" in cover or "date" in cover:
            info = []
            if "version" in cover:
                info.append(self.t["cover_version"].format(version=html.escape(cover["version"])))
            if "date" in cover:
                info.append(html.escape(cover["date"]))
            parts.append(f'<div class="cover-info">{" ".join(info)}</div>')

        meta_rows = []
        for field, label_key in _COVER_META_FIELDS:
            if field not in cover:
                continue
            value = cover[field]
            if isinstance(value, list):
                value = ", ".join(value)
            meta_rows.append((self.t[label_key], value))
        if meta_rows:
            parts.append('<div class="cover-meta"><table>')
            for label, value in meta_rows:
                parts.append(f"<tr><th>{label}:</th><td>{html.escape(value)}</td></tr>")
            parts.append("</table></div>")
        parts.append("</div>")
        return "\n".join(parts)

    def _revisions(self, doc: dict) -> str:
        parts = ['<div class="revision-page">']
        parts.append(f'<div class="header">{html.escape(doc["title"])}</div>')
        parts.append(f'<div class="revision-title">{self.t["revision_history"]}</div>')
        parts.append("<table>")
        parts.append(
            f'<tr><th style="width:60pt;">{self.t["version"]}</th>'
            f'<th style="width:80pt;">{self.t["date"]}</th>'
            f'<th style="width:80pt;">{self.t["author"]}</th>'
            f'<th>{self.t["summary"]}</th></tr>'
        )
        for rev in doc.get("revisions", []):
            parts.append(
                f'<tr><td style="text-align:center;">{html.escape(str(rev.get("version", "")))}</td>'
                f'<td style="text-align:center;">{html.escape(str(rev.get("date", "")))}</td>'
                f'<td>{html.escape(str(rev.get("author", "")))}</td>'
                f'<td>{html.escape(str(rev.get("summary", "")))}</td></tr>'
            )
        parts.append("</table></div>")
        return "\n".join(parts)

    def _section(self, section: dict, numbers: List[int]) -> str:
        number_str = ".".join(map(str, numbers))
        parts = []
        if section.get("page_break_before"):
            parts.append('<div class="page-break"></div>')
        h = min(len(numbers), 3)
        parts.append(f"<h{h}>{number_str}. {self._multiline(section.get('title', ''))}</h{h}>")
        for block in section.get("blocks", []) or []:
            parts.append(self._block(block, number_str))
        for i, child in enumerate(section.get("sections", []) or [], 1):
            parts.append(self._section(child, numbers + [i]))
        return "\n".join(parts)

    def _block(self, block: dict, section_num: str) -> str:
        renderers = {
            "text": lambda: self._text(block),
            "table": lambda: self._table(block, section_num),
            "list": lambda: self._list(block),
            "figure": lambda: self._figure(block),
            "image": lambda: self._image(block),
            "code": lambda: self._code(block),
            "note": lambda: self._note(block),
        }
        renderer = renderers.get(block.get("type"))
        return renderer() if renderer else f"<!-- unknown block type: {block.get('type')} -->"

    def _text(self, block: dict) -> str:
        paragraphs = block.get("content", "").strip().split("\n\n")
        return "\n".join(f"<p>{self._multiline(p.strip())}</p>" for p in paragraphs)

    def _table(self, block: dict, section_num: str) -> str:
        self.table_count += 1
        header = block.get("header", [])
        row_header_cols = block.get("row_header_cols", 0)

        caption_text = f'{self.t["table"]} {section_num}-{self.table_count}'
        if block.get("caption"):
            caption_text += f' {html.escape(block["caption"])}'
        parts = [f'<div class="table-caption">{caption_text}</div>', "<table>"]

        header_rows = header if header and isinstance(header[0], list) else [header]
        for hrow in header_rows:
            parts.append("<tr>")
            for cell in hrow:
                value, colspan, rowspan = _cell_parts(cell)
                attrs = ""
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                parts.append(f"<th{attrs}>{html.escape(_cell_text(value))}</th>")
            parts.append("</tr>")

        # rowspan continuation tracking: column index -> remaining covered rows
        rowspan_remaining: dict = {}
        for row in block.get("rows", []):
            parts.append("<tr>")
            col_idx = 0
            cell_idx = 0
            while cell_idx < len(row) or col_idx in rowspan_remaining:
                while col_idx in rowspan_remaining and rowspan_remaining[col_idx] > 0:
                    rowspan_remaining[col_idx] -= 1
                    if rowspan_remaining[col_idx] == 0:
                        del rowspan_remaining[col_idx]
                    col_idx += 1
                if cell_idx >= len(row):
                    break
                cell = row[cell_idx]
                value, colspan, rowspan = _cell_parts(cell)
                is_row_header = col_idx < row_header_cols
                attrs = ""
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                    rowspan_remaining[col_idx] = rowspan - 1
                if is_row_header:
                    parts.append(f'<th class="row-header"{attrs}>{self._multiline(_cell_text(value))}</th>')
                else:
                    parts.append(f"<td{attrs}>{self._multiline(_cell_text(value))}</td>")
                col_idx += colspan
                cell_idx += 1
            parts.append("</tr>")
        parts.append("</table>")
        return "\n".join(parts)

    def _list(self, block: dict) -> str:
        tag = "ol" if block.get("style") == "number" else "ul"
        return f"<{tag}>\n{self._list_items(block.get('items', []), tag)}\n</{tag}>"

    def _list_items(self, items: list, tag: str) -> str:
        parts = []
        for item in items:
            if isinstance(item, dict):
                text = html.escape(item.get("text", ""))
                children = item.get("children") or []
                if children:
                    parts.append(f"<li>{text}\n<{tag}>\n{self._list_items(children, tag)}\n</{tag}>\n</li>")
                else:
                    parts.append(f"<li>{text}</li>")
            else:
                parts.append(f"<li>{html.escape(str(item))}</li>")
        return "\n".join(parts)

    def _figure(self, block: dict) -> str:
        self.figure_count += 1
        caption_text = f'{self.t["figure"]} {self.figure_count}'
        if block.get("caption"):
            caption_text += f' {html.escape(block["caption"])}'
        return (
            '<div class="mermaid-container">\n'
            f'<div class="mermaid">\n{block.get("code", "")}\n</div>\n'
            f'<div class="figure-caption">{caption_text}</div>\n'
            "</div>"
        )

    def _image(self, block: dict) -> str:
        self.figure_count += 1
        caption = block.get("caption", "")
        alt = block.get("alt", caption)
        align = block.get("align", "left")
        align_class = f" align-{align}" if align != "left" else ""

        caption_text = f'{self.t["figure"]} {self.figure_count}'
        if caption:
            caption_text += f" {html.escape(caption)}"

        data_uri = self._resolve_image(block.get("path", ""))
        if data_uri:
            style = f' style="width:{html.escape(block["width"])};"' if block.get("width") else ""
            img = f'<img src="{data_uri}" alt="{html.escape(alt)}"{style}>'
        else:
            img = f'<span class="image-missing">{self.t["image_missing"]}</span>'

        return (
            f'<div class="image-container{align_class}">\n{img}\n'
            f'<div class="figure-caption">{caption_text}</div>\n</div>'
        )

    def _resolve_image(self, path: str) -> Optional[str]:
        """Resolve a workspace-relative path against the asset snapshot as a data URI."""
        if not self.assets_dir or not path:
            return None
        normalized = os.path.normpath(path)
        if normalized.startswith("..") or os.path.isabs(normalized):
            return None
        full = os.path.join(self.assets_dir, normalized)
        if not os.path.isfile(full):
            return None
        mime = mimetypes.guess_type(full)[0] or "application/octet-stream"
        with open(full, "rb") as f:
            encoded = base64.b64encode(f.read()).decode("ascii")
        return f"data:{mime};base64,{encoded}"

    def _code(self, block: dict) -> str:
        parts = []
        if block.get("caption"):
            parts.append(f'<div class="code-caption">{html.escape(block["caption"])}</div>')
        parts.append(f'<pre class="code-block"><code>{html.escape(block.get("content", ""))}</code></pre>')
        return "\n".join(parts)

    def _note(self, block: dict) -> str:
        style = block.get("style", "info")
        return f'<div class="note note-{style}">{self._multiline(block.get("content", "").strip())}</div>'


# ── Excel ──────────────────────────────────────────────────────────────────────

_CONTENT_COLS = 10  # blocks span columns A..J


class ExcelDocumentRenderer:
    """Renders a parsed document dict to an .xlsx workbook (bytes)."""

    def __init__(self, assets_dir: Optional[str] = None, generated_at: Optional[datetime] = None):
        self.assets_dir = assets_dir
        self.generated_at = generated_at
        self.table_count = 0
        self.figure_count = 0
        self.t = I18N["ja"]

    def render(self, doc: dict) -> bytes:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        self.table_count = 0
        self.figure_count = 0
        self.t = I18N.get(doc.get("language", "ja"), I18N["ja"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Document"
        for c in range(1, _CONTENT_COLS + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        self.row = 1

        self._cover(ws, doc)
        if doc.get("revisions"):
            self._page_break(ws)
            self._revisions(ws, doc)
        self._page_break(ws)
        for i, section in enumerate(doc.get("sections", []), 1):
            self._section(ws, section, [i])

        self.row += 1
        stamp = (self.generated_at or datetime.now()).strftime("%Y-%m-%d %H:%M")
        self._write_merged(ws, f'{self.t["generated_at"]}: {stamp}', italic=True, color="666666")

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── low-level helpers ──

    def _write_merged(self, ws, text, *, bold=False, italic=False, size=None,
                      fill=None, color=None, align="left", mono=False):
        """One logical line spanning the whole content width."""
        from openpyxl.styles import Alignment, Font, PatternFill

        cell = ws.cell(row=self.row, column=1, value=text)
        cell.font = Font(bold=bold, italic=italic, size=size or 11,
                         color=color or "000000", name="MS Gothic" if mono else None)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=align)
        if fill:
            f = PatternFill("solid", fgColor=fill)
            for c in range(1, _CONTENT_COLS + 1):
                ws.cell(row=self.row, column=c).fill = f
        ws.merge_cells(start_row=self.row, start_column=1,
                       end_row=self.row, end_column=_CONTENT_COLS)
        lines = str(text).count("\n") + 1
        if lines > 1:
            ws.row_dimensions[self.row].height = 14 * lines
        self.row += 1

    def _page_break(self, ws):
        from openpyxl.worksheet.pagebreak import Break

        if self.row > 1:
            ws.row_breaks.append(Break(id=self.row - 1))

    # ── document parts ──

    def _cover(self, ws, doc):
        cover = doc.get("cover", {})
        self.row += 2
        self._write_merged(ws, doc.get("title", ""), bold=True, size=20,
                           color="00876C", align="center")
        if "subtitle" in cover:
            self._write_merged(ws, cover["subtitle"], size=14, align="center")
        info = []
        if "version" in cover:
            info.append(self.t["cover_version"].format(version=cover["version"]))
        if "date" in cover:
            info.append(cover["date"])
        if info:
            self._write_merged(ws, "  ".join(info), size=12, align="center")
        self.row += 1
        for field, label_key in _COVER_META_FIELDS:
            if field not in cover:
                continue
            value = cover[field]
            if isinstance(value, list):
                value = ", ".join(value)
            self._write_merged(ws, f"{self.t[label_key]}: {value}", align="center")

    def _revisions(self, ws, doc):
        self._write_merged(ws, self.t["revision_history"], bold=True, size=14, align="center")
        header = [self.t["version"], self.t["date"], self.t["author"], self.t["summary"]]
        rows = [
            [str(r.get("version", "")), str(r.get("date", "")),
             str(r.get("author", "")), str(r.get("summary", ""))]
            for r in doc.get("revisions", [])
        ]
        self._grid(ws, [header], rows, row_header_cols=0)

    def _section(self, ws, section, numbers):
        if section.get("page_break_before"):
            self._page_break(ws)
        number_str = ".".join(map(str, numbers))
        level = min(len(numbers), 3)
        title = f"{number_str}. {section.get('title', '')}"
        if level == 1:
            self._write_merged(ws, title, bold=True, size=14, fill="00876C", color="FFFFFF")
        elif level == 2:
            self._write_merged(ws, title, bold=True, size=12, fill="F0F0F0")
        else:
            self._write_merged(ws, title, bold=True, size=11)
        for block in section.get("blocks", []) or []:
            self._block(ws, block, number_str)
        for i, child in enumerate(section.get("sections", []) or [], 1):
            self._section(ws, child, numbers + [i])

    def _block(self, ws, block, section_num):
        kind = block.get("type")
        if kind == "text":
            for para in block.get("content", "").strip().split("\n\n"):
                self._write_merged(ws, para.strip())
        elif kind == "table":
            self._table(ws, block, section_num)
        elif kind == "list":
            self._list_items(ws, block.get("items", []), block.get("style"), 0)
        elif kind == "figure":
            self.figure_count += 1
            self._write_merged(ws, block.get("code", "").rstrip(), fill="F5F5F5", mono=True)
            self._caption(ws, self.t["figure"], self.figure_count, block.get("caption"))
        elif kind == "image":
            self._image(ws, block)
        elif kind == "code":
            if block.get("caption"):
                self._write_merged(ws, block["caption"])
            self._write_merged(ws, block.get("content", "").rstrip(), fill="F5F5F5", mono=True)
        elif kind == "note":
            fills = {"info": "E8F4FC", "warning": "FFF8E6", "important": "FFEEEE"}
            self._write_merged(ws, block.get("content", "").strip(),
                               fill=fills.get(block.get("style", "info"), "F9F9F9"))

    def _caption(self, ws, label, count, caption, section_num=None):
        text = f"{label} {section_num}-{count}" if section_num else f"{label} {count}"
        if caption:
            text += f" {caption}"
        self._write_merged(ws, text, bold=True, size=10)

    def _table(self, ws, block, section_num):
        self.table_count += 1
        self._caption(ws, self.t["table"], self.table_count, block.get("caption"), section_num)
        header = block.get("header", [])
        header_rows = header if header and isinstance(header[0], list) else [header]
        self._grid(ws, header_rows, block.get("rows", []),
                   block.get("row_header_cols", 0))

    def _grid(self, ws, header_rows, data_rows, row_header_cols):
        """Write a bordered table grid with merges, mirroring the HTML algorithm."""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="D9D9D9")

        def put(row, col, value, *, is_header, colspan=1, rowspan=1):
            cell = ws.cell(row=row, column=col, value=_cell_text(value))
            cell.alignment = Alignment(
                wrap_text=True, vertical="top",
                horizontal="center" if is_header else "left",
            )
            cell.font = Font(size=10)
            for r in range(row, row + rowspan):
                for c in range(col, col + colspan):
                    ws.cell(row=r, column=c).border = border
                    if is_header:
                        ws.cell(row=r, column=c).fill = header_fill
            if colspan > 1 or rowspan > 1:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row + rowspan - 1, end_column=col + colspan - 1)

        for hrow in header_rows:
            col = 1
            for cell in hrow:
                value, colspan, rowspan = _cell_parts(cell)
                put(self.row, col, value, is_header=True, colspan=colspan, rowspan=rowspan)
                col += colspan
            self.row += 1

        rowspan_remaining: dict = {}
        for row_cells in data_rows:
            col_idx = 0  # 0-based grid column, as in the HTML renderer
            cell_idx = 0
            while cell_idx < len(row_cells) or col_idx in rowspan_remaining:
                while col_idx in rowspan_remaining and rowspan_remaining[col_idx] > 0:
                    rowspan_remaining[col_idx] -= 1
                    if rowspan_remaining[col_idx] == 0:
                        del rowspan_remaining[col_idx]
                    col_idx += 1
                if cell_idx >= len(row_cells):
                    break
                value, colspan, rowspan = _cell_parts(row_cells[cell_idx])
                if rowspan > 1:
                    rowspan_remaining[col_idx] = rowspan - 1
                put(self.row, col_idx + 1, value,
                    is_header=col_idx < row_header_cols, colspan=colspan, rowspan=rowspan)
                col_idx += colspan
                cell_idx += 1
            self.row += 1
        self.row += 1  # spacing after the table

    def _list_items(self, ws, items, style, depth):
        from openpyxl.styles import Alignment, Font

        for n, item in enumerate(items, 1):
            if isinstance(item, dict):
                text, children = item.get("text", ""), item.get("children") or []
            else:
                text, children = str(item), []
            prefix = f"{n}." if style == "number" else "-"
            cell = ws.cell(row=self.row, column=1, value=f"{'    ' * depth}{prefix} {text}")
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=self.row, start_column=1,
                           end_row=self.row, end_column=_CONTENT_COLS)
            self.row += 1
            if children:
                self._list_items(ws, children, style, depth + 1)

    def _image(self, ws, block):
        self.figure_count += 1
        path = self._resolve_image_path(block.get("path", ""))
        if path:
            try:
                from openpyxl.drawing.image import Image as XlImage

                img = XlImage(path)
                # Scale down to roughly the content width if wider.
                max_px = 640
                if img.width and img.width > max_px:
                    ratio = max_px / img.width
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                ws.add_image(img, f"A{self.row}")
                self.row += max(2, int((img.height or 100) / 18) + 1)
            except Exception:
                self._write_merged(ws, self.t["image_missing"], color="666666")
        else:
            self._write_merged(ws, self.t["image_missing"], color="666666")
        self._caption(ws, self.t["figure"], self.figure_count, block.get("caption"))

    def _resolve_image_path(self, path: str) -> Optional[str]:
        if not self.assets_dir or not path:
            return None
        normalized = os.path.normpath(path)
        if normalized.startswith("..") or os.path.isabs(normalized):
            return None
        full = os.path.join(self.assets_dir, normalized)
        return full if os.path.isfile(full) else None


def render_document(doc: dict, fmt: str, *, assets_dir: Optional[str] = None,
                    generated_at: Optional[datetime] = None):
    """Render a parsed document. Returns str for 'html', bytes for 'excel'."""
    if fmt == "html":
        return HtmlDocumentRenderer(assets_dir, generated_at).render(doc)
    if fmt == "excel":
        return ExcelDocumentRenderer(assets_dir, generated_at).render(doc)
    raise ValueError(f"Unsupported format: {fmt}")
